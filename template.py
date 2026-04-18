"""
Template generation module.

The system owns the template shape — users must only download, fill, and
re-upload. No fragile formulas live in the user sheets; all logic runs in
the software.
"""
from __future__ import annotations

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from .config import (
        APP_NAME, APP_TAGLINE, TEMPLATE_VERSION, LOGIC_VERSION,
        SHEET_README, SHEET_ICG, SHEET_DMM, SHEET_GPIS_SUPPLY, SHEET_GPIS_DEMAND,
        SHEET_LOOKUPS,
        ICG_COLUMNS, DMM_COLUMNS, GPIS_SUPPLY_COLUMNS, GPIS_DEMAND_COLUMNS,
        DOMAINS, GEOGRAPHIES,
    )
except ImportError:
    from config import (
        APP_NAME, APP_TAGLINE, TEMPLATE_VERSION, LOGIC_VERSION,
        SHEET_README, SHEET_ICG, SHEET_DMM, SHEET_GPIS_SUPPLY, SHEET_GPIS_DEMAND,
        SHEET_LOOKUPS,
        ICG_COLUMNS, DMM_COLUMNS, GPIS_SUPPLY_COLUMNS, GPIS_DEMAND_COLUMNS,
        DOMAINS, GEOGRAPHIES,
    )

# Visual tokens
INK = "0F172A"
ACCENT = "1E3A8A"
SOFT = "E0E7FF"
MUTED = "64748B"
DIVIDER = "CBD5E1"


def _header_style(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=DIVIDER),
        right=Side(style="thin", color=DIVIDER),
        top=Side(style="thin", color=DIVIDER),
        bottom=Side(style="thin", color=DIVIDER),
    )


def _write_headers(ws, columns):
    for idx, (label, _spec) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(idx)].width = max(18, len(label) + 4)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _apply_dropdowns(ws, columns, n_rows: int = 500):
    """Attach data-validation dropdowns for closed-list columns."""
    for idx, (label, spec) in enumerate(columns, start=1):
        col_letter = get_column_letter(idx)
        rng = f"{col_letter}2:{col_letter}{n_rows + 1}"
        if isinstance(spec, list):
            formula = '"' + ",".join(spec) + '"'
            dv = DataValidation(
                type="list", formula1=formula, allow_blank=True, showDropDown=False
            )
            dv.error = f"Choose one of: {', '.join(spec)}"
            dv.errorTitle = "Invalid value"
            dv.prompt = f"Allowed: {', '.join(spec)}"
            dv.promptTitle = label
            ws.add_data_validation(dv)
            dv.add(rng)
        elif spec == "domain":
            dv = DataValidation(
                type="list",
                formula1=f"={SHEET_LOOKUPS}!$A$2:$A${len(DOMAINS) + 1}",
                allow_blank=True,
                showDropDown=False,
            )
            dv.prompt = "Pick a domain from the master list."
            dv.promptTitle = label
            ws.add_data_validation(dv)
            dv.add(rng)
        elif spec == "geography":
            dv = DataValidation(
                type="list",
                formula1=f"={SHEET_LOOKUPS}!$B$2:$B${len(GEOGRAPHIES) + 1}",
                allow_blank=True,
                showDropDown=False,
            )
            dv.prompt = "Pick a geography from the master list."
            dv.promptTitle = label
            ws.add_data_validation(dv)
            dv.add(rng)
        elif spec == "int":
            dv = DataValidation(
                type="whole",
                operator="greaterThanOrEqual",
                formula1=0,
                allow_blank=True,
            )
            dv.prompt = "Enter a whole number (≥ 0)."
            dv.promptTitle = label
            ws.add_data_validation(dv)
            dv.add(rng)


def _build_readme(ws):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    header = ws.cell(row=1, column=1, value=f"{APP_NAME} — {APP_TAGLINE}")
    header.font = Font(size=18, bold=True, color=ACCENT)
    ws.merge_cells("A1:B1")

    sub = ws.cell(
        row=2,
        column=1,
        value=(
            f"Template v{TEMPLATE_VERSION} · Logic v{LOGIC_VERSION} · Generated "
            f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ),
    )
    sub.font = Font(size=10, italic=True, color=MUTED)
    ws.merge_cells("A2:B2")

    rows = [
        ("", ""),
        (
            "What this is",
            "A multi-layer institutional health diagnostics engine. It evaluates "
            "institutional continuity (ICG), programme vitality (DMM), and market "
            "alignment (GPIS).",
        ),
        ("What this is not", "Not a ranking system. Not a compliance dashboard. Not a data warehouse."),
        ("", ""),
        (
            "How to use",
            "1. Fill ICG_Input, DMM_Input, GPIS_Supply, GPIS_Demand. 2. Use dropdowns "
            "wherever available. 3. Save and upload the same workbook back into the "
            "EduPulse app.",
        ),
        ("Do not", "Do not rename sheets. Do not rename header columns. Do not add computed formulas — the engine does all scoring."),
        ("", ""),
        (SHEET_ICG, "One row per faculty member. Measures continuity, dependency, succession, and attrition exposure."),
        (SHEET_DMM, "One row per programme. Measures programme vitality, value, adaptation, alignment, contribution, and teaching effectiveness."),
        (SHEET_GPIS_SUPPLY, "One row per (programme × domain × geography). Lists what the institution supplies."),
        (SHEET_GPIS_DEMAND, "One row per (domain × geography). Describes market demand signals."),
        ("", ""),
        ("Confidence", "The engine computes a coverage score per framework; missing rows/values reduce confidence but do not block analysis."),
        ("Privacy", "All processing happens in the hosted app session. No data is persisted server-side beyond your active session unless you export it."),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        kc = ws.cell(row=i, column=1, value=k)
        vc = ws.cell(row=i, column=2, value=v)
        kc.font = Font(bold=True, color=INK)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        if not k and not v:
            continue
        if k in {SHEET_ICG, SHEET_DMM, SHEET_GPIS_SUPPLY, SHEET_GPIS_DEMAND}:
            kc.fill = PatternFill("solid", fgColor=SOFT)


def _build_lookups(ws):
    ws.cell(row=1, column=1, value="Domains").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Geographies").font = Font(bold=True)
    for i, d in enumerate(DOMAINS, start=2):
        ws.cell(row=i, column=1, value=d)
    for i, g in enumerate(GEOGRAPHIES, start=2):
        ws.cell(row=i, column=2, value=g)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 24
    ws.sheet_state = "hidden"


def build_template_workbook() -> Workbook:
    """Return an openpyxl Workbook containing the master template."""
    wb = Workbook()

    ws_readme = wb.active
    ws_readme.title = SHEET_README
    _build_readme(ws_readme)

    for sheet, cols in [
        (SHEET_ICG, ICG_COLUMNS),
        (SHEET_DMM, DMM_COLUMNS),
        (SHEET_GPIS_SUPPLY, GPIS_SUPPLY_COLUMNS),
        (SHEET_GPIS_DEMAND, GPIS_DEMAND_COLUMNS),
    ]:
        ws = wb.create_sheet(sheet)
        _write_headers(ws, cols)

    ws_lk = wb.create_sheet(SHEET_LOOKUPS)
    _build_lookups(ws_lk)

    for sheet, cols in [
        (SHEET_ICG, ICG_COLUMNS),
        (SHEET_DMM, DMM_COLUMNS),
        (SHEET_GPIS_SUPPLY, GPIS_SUPPLY_COLUMNS),
        (SHEET_GPIS_DEMAND, GPIS_DEMAND_COLUMNS),
    ]:
        _apply_dropdowns(wb[sheet], cols)

    return wb


def template_bytes() -> bytes:
    """Return the master template as bytes suitable for download."""
    wb = build_template_workbook()
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
