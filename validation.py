"""
Upload validation engine.

Reads an uploaded workbook against the locked EduPulse template schema and
returns normalized DataFrames + a validation report.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

from .config import (
    SHEET_ICG, SHEET_DMM, SHEET_GPIS_SUPPLY, SHEET_GPIS_DEMAND,
    REQUIRED_SHEETS, DOMAINS, GEOGRAPHIES,
    ICG_COLUMNS, DMM_COLUMNS, GPIS_SUPPLY_COLUMNS, GPIS_DEMAND_COLUMNS,
)


@dataclass
class ValidationReport:
    ok: bool = True
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    sheet_row_counts: Dict[str, int] = field(default_factory=dict)

    def fail(self, msg: str):
        self.ok = False
        self.errors.append(msg)

    def warn(self, msg: str):
        self.warnings.append(msg)


SHEET_SPECS = {
    SHEET_ICG: ICG_COLUMNS,
    SHEET_DMM: DMM_COLUMNS,
    SHEET_GPIS_SUPPLY: GPIS_SUPPLY_COLUMNS,
    SHEET_GPIS_DEMAND: GPIS_DEMAND_COLUMNS,
}


def _read_sheet(xls_bytes: bytes, sheet: str) -> pd.DataFrame:
    return pd.read_excel(BytesIO(xls_bytes), sheet_name=sheet, engine="openpyxl")


def _normalise_cell(v):
    if isinstance(v, str):
        return v.strip()
    return v


def _validate_columns(df: pd.DataFrame, expected: list, sheet: str, report: ValidationReport) -> pd.DataFrame:
    expected_labels = [label for label, _ in expected]
    missing = [c for c in expected_labels if c not in df.columns]
    if missing:
        report.fail(f"[{sheet}] Missing columns: {', '.join(missing)}")
        return df
    # re-order to expected order and drop extras
    df = df[expected_labels].copy()
    # strip whitespace for string cells
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(_normalise_cell)
    return df


def _validate_values(df: pd.DataFrame, expected: list, sheet: str, report: ValidationReport):
    for label, spec in expected:
        series = df[label]
        if isinstance(spec, list):
            allowed = set(spec)
            mask = series.notna() & ~series.isin(allowed)
            if mask.any():
                bad_rows = df.index[mask].tolist()
                sample = series[mask].unique().tolist()[:5]
                report.warn(
                    f"[{sheet}] '{label}' has {int(mask.sum())} value(s) outside the allowed list "
                    f"(rows {bad_rows[:5]}…; e.g. {sample}). These rows will be treated as missing."
                )
                # blank them — downstream treats NaN as missing
                df.loc[mask, label] = None
        elif spec == "int":
            # coerce to numeric
            df[label] = pd.to_numeric(series, errors="coerce")
            neg = df[label] < 0
            if neg.any():
                report.warn(f"[{sheet}] '{label}' has negative values; they will be set to 0.")
                df.loc[neg, label] = 0
        elif spec == "domain":
            allowed = set(DOMAINS)
            mask = series.notna() & ~series.isin(allowed)
            if mask.any():
                sample = series[mask].unique().tolist()[:5]
                report.warn(
                    f"[{sheet}] '{label}' contains values outside the master domain list: {sample}. "
                    f"These rows will be treated as missing for domain-match logic."
                )
                df.loc[mask, label] = None
        elif spec == "geography":
            allowed = set(GEOGRAPHIES)
            mask = series.notna() & ~series.isin(allowed)
            if mask.any():
                sample = series[mask].unique().tolist()[:5]
                report.warn(
                    f"[{sheet}] '{label}' contains values outside the master geography list: {sample}. "
                    f"These rows will be treated as missing for geography-match logic."
                )
                df.loc[mask, label] = None


def validate_workbook(xls_bytes: bytes):
    """Validate a workbook and return (dfs: dict, report: ValidationReport)."""
    report = ValidationReport()
    dfs: Dict[str, pd.DataFrame] = {}

    # ensure sheet presence
    try:
        wb = load_workbook(BytesIO(xls_bytes), read_only=True, data_only=True)
        present = set(wb.sheetnames)
    except Exception as e:
        report.fail(f"Could not open workbook: {e}")
        return dfs, report

    missing_sheets = [s for s in REQUIRED_SHEETS if s not in present]
    if missing_sheets:
        report.fail(f"Missing required sheets: {', '.join(missing_sheets)}. "
                    f"Please use the official EduPulse template.")
        return dfs, report

    for sheet, spec in SHEET_SPECS.items():
        try:
            df = _read_sheet(xls_bytes, sheet)
        except Exception as e:
            report.fail(f"[{sheet}] Could not read sheet: {e}")
            continue
        # drop rows that are entirely empty
        df = df.dropna(how="all").reset_index(drop=True)
        df = _validate_columns(df, spec, sheet, report)
        if report.ok:
            _validate_values(df, spec, sheet, report)
        report.sheet_row_counts[sheet] = len(df)
        dfs[sheet] = df

    # required-presence warnings
    if dfs.get(SHEET_ICG, pd.DataFrame()).empty:
        report.warn(f"[{SHEET_ICG}] is empty — ICG analysis will be skipped.")
    if dfs.get(SHEET_DMM, pd.DataFrame()).empty:
        report.warn(f"[{SHEET_DMM}] is empty — DMM analysis will be skipped.")
    if dfs.get(SHEET_GPIS_SUPPLY, pd.DataFrame()).empty:
        report.warn(f"[{SHEET_GPIS_SUPPLY}] is empty — GPIS analysis will be skipped.")
    if dfs.get(SHEET_GPIS_DEMAND, pd.DataFrame()).empty:
        report.warn(f"[{SHEET_GPIS_DEMAND}] is empty — GPIS demand matching cannot proceed.")

    return dfs, report
